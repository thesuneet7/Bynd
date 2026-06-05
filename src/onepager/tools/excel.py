"""Local Excel/CSV extraction into markdown-ish sheet chunks."""
from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path


@dataclass
class ParsedSheet:
    sheet: str
    markdown: str


def parse_spreadsheet(path: Path, *, max_rows: int = 120, max_cols: int = 30) -> list[ParsedSheet]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _parse_csv(path, max_rows=max_rows, max_cols=max_cols)
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return _parse_xlsx(path, max_rows=max_rows, max_cols=max_cols)
    return []


def _parse_csv(path: Path, *, max_rows: int, max_cols: int) -> list[ParsedSheet]:
    try:
        rows: list[list[str]] = []
        with open(path, newline="", encoding="utf-8-sig") as f:
            for i, row in enumerate(csv.reader(f)):
                if i >= max_rows:
                    break
                rows.append([str(c).strip() for c in row[:max_cols]])
        text = _rows_to_markdown(rows)
        return [ParsedSheet(sheet=path.stem, markdown=text)] if text.strip() else []
    except Exception:
        return []


def _parse_xlsx(path: Path, *, max_rows: int, max_cols: int) -> list[ParsedSheet]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return []
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        sheets: list[ParsedSheet] = []
        for ws in wb.worksheets:
            rows: list[list[str]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_rows:
                    break
                vals = ["" if v is None else str(v).strip() for v in row[:max_cols]]
                if any(vals):
                    rows.append(vals)
            text = _rows_to_markdown(rows)
            if text.strip():
                sheets.append(ParsedSheet(sheet=ws.title, markdown=text))
        return sheets
    except Exception:
        return []


def _rows_to_markdown(rows: list[list[str]]) -> str:
    lines: list[str] = []
    for idx, row in enumerate(rows, start=1):
        if not any(cell.strip() for cell in row):
            continue
        lines.append(f"Row {idx}: " + " | ".join(cell for cell in row))
    return "\n".join(lines)
